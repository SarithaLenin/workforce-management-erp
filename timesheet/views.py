from django.shortcuts import render,redirect,get_object_or_404
import calendar
from datetime import date
from companies.models import Company
from clients.models import ProjectSite, EmployeeAssignment
from attendance.models import DailyAttendance
from django.contrib import messages
from .models import *
from django.db.models import Q
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from invoice.models import Invoice
from django.contrib.auth.decorators import login_required
from accounts.decorators import role_required
from django.views.decorators.http import require_POST



# Create your views here.

@login_required
@role_required(["admin", "hr", "accountant"])
def timesheet_generate(request):

    companies = Company.objects.filter(is_active=True).order_by('name')
    projects = ProjectSite.objects.filter(is_active=True).order_by('name')

    selected_company = None
    selected_project = None
    selected_month = None
    selected_year = None
    existing_timesheet= None
    month_name = None
    start_date = None
    end_date = None

    days = []
    timesheet_rows = []
    daily_totals = {}
    grand_total_hours = 0

    if request.GET.get('company') and request.GET.get('project') and request.GET.get('month') and request.GET.get('year'):

        selected_company = Company.objects.get(id=request.GET.get('company'))
        selected_project = ProjectSite.objects.get(id=request.GET.get('project'))
        selected_month = int(request.GET.get('month'))
        selected_year = int(request.GET.get('year'))

        existing_timesheet = MonthlyTimesheet.objects.filter(
            project_site=selected_project,
            month=selected_month,
            year=selected_year
        ).first()

        total_days = calendar.monthrange(selected_year, selected_month)[1]
        days = list(range(1, total_days + 1))

        month_name = calendar.month_name[selected_month]
        start_date = date(selected_year, selected_month, 1)
        end_date = date(selected_year, selected_month, total_days)

        assignments = EmployeeAssignment.objects.select_related(
            'employee',
            'designation'
        ).filter(
            project_site=selected_project
        ).order_by(
            'employee__first_name',
            'employee__last_name'
        )

        attendances = DailyAttendance.objects.filter(
            project_site=selected_project,
            attendance_date__year=selected_year,
            attendance_date__month=selected_month
        )

        attendance_map = {}

        for att in attendances:
            attendance_map[(att.employee_id, att.attendance_date.day)] = att

        for assignment in assignments:

            row_days = []
            total_hours = 0

            for day in days:

                current_date = date(selected_year, selected_month, day)
                att = attendance_map.get((assignment.employee_id, day))

                value = ""

                if att:
                    if att.status == 'PRESENT':
                        value = att.total_hours
                        total_hours += att.total_hours
                        daily_totals[day] = daily_totals.get(day, 0) + att.total_hours

                    elif att.status == 'ABSENT':
                        value = "A"

                    elif att.status == 'LEAVE':
                        value = "L"

                    elif att.status == 'HOLIDAY':
                        value = "H"

                else:
                    if current_date.weekday() == 6:
                        value = "S"

                row_days.append(value)

            timesheet_rows.append({
                'employee': assignment.employee,
                'designation': assignment.designation,
                'days': row_days,
                'total_hours': total_hours,
            })

        if request.GET.get('save') == '1':

            timesheet, created = MonthlyTimesheet.objects.get_or_create(
                project_site=selected_project,
                month=selected_month,
                year=selected_year,
                defaults={
                    'company': selected_company,
                    'status': 'DRAFT'
                }
            )

            if created:

                for row in timesheet_rows:

                    day_values = {}

                    for index, value in enumerate(row['days'], start=1):
                        day_values[f'day_{index}'] = str(value)

                    MonthlyTimesheetRow.objects.create(
                    timesheet=timesheet,
                    employee=row["employee"],
                    employee_name=f"{row['employee'].first_name} {row['employee'].last_name}".strip(),
                    designation=row["designation"].name if row["designation"] else "",
                    total_hours=row["total_hours"],
                    **day_values,
                   )

                messages.success(request, 'Timesheet saved successfully.')

            else:
                messages.warning(request, 'Timesheet already exists for this project and month.')

            return redirect('timesheet_list')

    grand_total_hours = sum(daily_totals.values())

    context = {
        'companies': companies,
        'projects': projects,
        'selected_company': selected_company,
        'selected_project': selected_project,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'existing_timesheet':  existing_timesheet,
        'days': days,
        'timesheet_rows': timesheet_rows,
        'daily_totals': daily_totals,
        'month_name': month_name,
        'start_date': start_date,
        'end_date': end_date,
        'grand_total_hours': grand_total_hours,
        
    }

    return render(request, 'timesheet/timesheet_generate.html', context)




@login_required
@role_required(["admin", "hr", "accountant"])
def timesheet_list(request):

    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')

    timesheets = MonthlyTimesheet.objects.select_related(
        'company',
        'project_site',
        'project_site__client'
    ).order_by('-year', '-month', '-id')

    if search:
        timesheets = timesheets.filter(
            Q(company__name__icontains=search) |
            Q(project_site__name__icontains=search) |
            Q(project_site__client__name__icontains=search)
        )

    if status:
        timesheets = timesheets.filter(status=status)

    if month:
        timesheets = timesheets.filter(month=month)

    if year:
        timesheets = timesheets.filter(year=year)

    return render(request, 'timesheet/timesheet_list.html', {
        'timesheets': timesheets,
        'search': search,
        'status': status,
        'month': month,
        'year': year,
    })




@login_required
@role_required(["admin", "hr", "accountant"])

def timesheet_detail(request, pk):

    timesheet = get_object_or_404(MonthlyTimesheet, pk=pk)

    total_days = calendar.monthrange(timesheet.year, timesheet.month)[1]
    days = list(range(1, total_days + 1))

    month_name = calendar.month_name[timesheet.month]
    start_date = date(timesheet.year, timesheet.month, 1)
    end_date = date(timesheet.year, timesheet.month, total_days)

    rows = []
    daily_totals = {day: 0 for day in days}
    grand_total_hours = 0

    for row in timesheet.rows.all().order_by('employee_name'):

        row_days = []

        for day in days:
            value = getattr(row, f'day_{day}')
            row_days.append(value)

            try:
                daily_totals[day] += float(value)
            except:
                pass

        grand_total_hours += float(row.total_hours)

        rows.append({
            'employee_name': row.employee_name,
            'designation': row.designation,
            'days': row_days,
            'total_hours': row.total_hours,
        })

    return render(request, 'timesheet/timesheet_detail.html', {
        'timesheet': timesheet,
        'days': days,
        'rows': rows,
        'daily_totals': daily_totals,
        'grand_total_hours': grand_total_hours,
        'month_name': month_name,
        'start_date': start_date,
        'end_date': end_date,
    })





@login_required
@role_required(["admin", "hr", "accountant"])

def timesheet_export_excel(request, pk):

    timesheet = get_object_or_404(MonthlyTimesheet, pk=pk)

    total_days = calendar.monthrange(timesheet.year, timesheet.month)[1]
    month_name = calendar.month_name[timesheet.month]

    start_date = date(timesheet.year, timesheet.month, 1)
    end_date = date(timesheet.year, timesheet.month, total_days)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sunday_fill = PatternFill("solid", fgColor="FFF2CC")
    total_fill = PatternFill("solid", fgColor="E2F0D9")

    max_col = 3 + total_days + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = f"Details of Manpower Hired from M/s : {timesheet.company.name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = f"Project Name : {timesheet.project_site.name}"
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    current_row = 3

    if timesheet.project_site.project_code:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
        ws[f"A{current_row}"] = f"Project Number : {timesheet.project_site.project_code}"
        ws[f"A{current_row}"].font = Font(bold=True)
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center")
        current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
    ws[f"A{current_row}"] = (
        f"{month_name.upper()}-{timesheet.year} "
        f"({start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')})"
    )
    ws[f"A{current_row}"].font = Font(bold=True)
    ws[f"A{current_row}"].alignment = Alignment(horizontal="center")

    table_start = current_row + 2

    headers = ["S.No", "Name", "Designation"]

    for day in range(1, total_days + 1):
        headers.append(f"{day:02d}")

    headers.append("Total Hours")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = header_fill

    row_no = table_start + 1
    daily_totals = {day: 0 for day in range(1, total_days + 1)}
    grand_total = 0

    for index, row in enumerate(timesheet.rows.all().order_by("employee_name"), start=1):

        ws.cell(row=row_no, column=1, value=index)
        ws.cell(row=row_no, column=2, value=row.employee_name)
        ws.cell(row=row_no, column=3, value=row.designation)

        ws.cell(row=row_no, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_no, column=2).alignment = Alignment(horizontal="left")
        ws.cell(row=row_no, column=3).alignment = Alignment(horizontal="center")

        for col in range(1, 4):
            ws.cell(row=row_no, column=col).border = border

        for day in range(1, total_days + 1):

            value = getattr(row, f"day_{day}")
            col = 3 + day

            display_value = value

            try:
                num = float(value)

                if num.is_integer():
                    display_value = int(num)
                else:
                    display_value = num

            except:
                display_value = value

            cell = ws.cell(
                row=row_no,
                column=col,
                value=display_value
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            current_date = date(timesheet.year, timesheet.month, day)

            if current_date.weekday() == 6:
                cell.fill = sunday_fill

            try:
                daily_totals[day] += float(value)
            except:
                pass

        total_col = 3 + total_days + 1
        ws.cell(row=row_no, column=total_col, value=float(row.total_hours))
        ws.cell(row=row_no, column=total_col).font = Font(bold=True)
        ws.cell(row=row_no, column=total_col).alignment = Alignment(horizontal="center")
        ws.cell(row=row_no, column=total_col).border = border

        grand_total += float(row.total_hours)

        row_no += 1

    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=3)
    ws.cell(row=row_no, column=1, value="Daily Total Hours")
    ws.cell(row=row_no, column=1).font = Font(bold=True)
    ws.cell(row=row_no, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row_no, column=1).fill = total_fill

    for col in range(1, 4):
        ws.cell(row=row_no, column=col).border = border
        ws.cell(row=row_no, column=col).fill = total_fill

    for day in range(1, total_days + 1):
        col = 3 + day
        cell = ws.cell(row=row_no, column=col, value=daily_totals[day])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell.fill = total_fill

    total_col = 3 + total_days + 1
    ws.cell(row=row_no, column=total_col, value=grand_total)
    ws.cell(row=row_no, column=total_col).font = Font(bold=True)
    ws.cell(row=row_no, column=total_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row_no, column=total_col).border = border
    ws.cell(row=row_no, column=total_col).fill = total_fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16

    for col in range(4, total_col):
        ws.column_dimensions[get_column_letter(col)].width = 5

    ws.column_dimensions[get_column_letter(total_col)].width = 12

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    filename = f"Timesheet_{timesheet.get_month_display()}_{timesheet.year}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response


@login_required
@role_required(["admin", "hr", "accountant"])
@require_POST
def timesheet_approve(request, pk):
    timesheet = get_object_or_404(
        MonthlyTimesheet,
        pk=pk,
    )

    if timesheet.status != "DRAFT":
        messages.error(
            request,
            "Only draft timesheets can be approved.",
        )
        return redirect(
            "timesheet_detail",
            pk=timesheet.pk,
        )

    timesheet.status = "APPROVED"
    timesheet.save(update_fields=["status"])

    messages.success(
        request,
        "Timesheet approved successfully.",
    )

    return redirect(
        "timesheet_detail",
        pk=timesheet.pk,
    )


@login_required
@role_required(["admin"])
@require_POST
def timesheet_reopen(request, pk):
    timesheet = get_object_or_404(
        MonthlyTimesheet,
        pk=pk,
    )

    if timesheet.status != "APPROVED":
        messages.error(
            request,
            "Only approved timesheets can be reopened.",
        )
        return redirect(
            "timesheet_detail",
            pk=timesheet.pk,
        )

    if Invoice.objects.filter(timesheet=timesheet).exists():
        messages.error(
            request,
            "This timesheet has an invoice. Manage the invoice first.",
        )
        return redirect(
            "timesheet_detail",
            pk=timesheet.pk,
        )

    timesheet.status = "DRAFT"
    timesheet.save(update_fields=["status"])

    messages.success(
        request,
        "Timesheet reopened successfully.",
    )

    return redirect(
        "timesheet_detail",
        pk=timesheet.pk,
    )


@login_required
@role_required(["admin", "hr", "accountant"])
def timesheet_delete(request, pk):
    timesheet = get_object_or_404(
        MonthlyTimesheet,
        pk=pk,
    )

    user_role = getattr(
        getattr(request.user, "userprofile", None),
        "role",
        None,
    )

    is_admin = (
        request.user.is_superuser
        or user_role == "admin"
    )

    if timesheet.status == "APPROVED" and not is_admin:
        messages.error(
            request,
            "Only an administrator can delete an approved timesheet.",
        )
        return redirect(
            "timesheet_detail",
            pk=timesheet.pk,
        )

    if Invoice.objects.filter(timesheet=timesheet).exists():
        messages.error(
            request,
            (
                "This timesheet has a generated invoice "
                "and cannot be deleted."
            ),
        )
        return redirect(
            "timesheet_detail",
            pk=timesheet.pk,
        )

    if request.method == "POST":
        timesheet.delete()

        messages.success(
            request,
            "Timesheet deleted successfully.",
        )
        return redirect("timesheet_list")

    return render(
        request,
        "timesheet/timesheet_confirm_delete.html",
        {"timesheet": timesheet},
    )



@login_required
@role_required(["admin", "hr", "accountant"])
def timesheet_print(request, pk):
    timesheet = get_object_or_404(
        MonthlyTimesheet,
        pk=pk,
    )

    total_days = calendar.monthrange(
        timesheet.year,
        timesheet.month,
    )[1]

    days = list(range(1, total_days + 1))

    month_name = calendar.month_name[timesheet.month]

    start_date = date(
        timesheet.year,
        timesheet.month,
        1,
    )

    end_date = date(
        timesheet.year,
        timesheet.month,
        total_days,
    )

    rows = []
    daily_totals = {
        day: 0
        for day in days
    }
    grand_total_hours = 0

    for row in timesheet.rows.all().order_by(
        "employee_name"
    ):
        row_days = []

        for day in days:
            value = getattr(row, f"day_{day}")
            row_days.append(value)

            try:
                daily_totals[day] += float(value)
            except (TypeError, ValueError):
                pass

        grand_total_hours += float(row.total_hours)

        rows.append({
            "employee_name": row.employee_name,
            "designation": row.designation,
            "days": row_days,
            "total_hours": row.total_hours,
        })

    return render(
        request,
        "timesheet/timesheet_print.html",
        {
            "timesheet": timesheet,
            "days": days,
            "rows": rows,
            "daily_totals": daily_totals,
            "grand_total_hours": grand_total_hours,
            "month_name": month_name,
            "start_date": start_date,
            "end_date": end_date,
        },
    )