import os

from openpyxl import Workbook
from openpyxl.styles import ( Font, Alignment, Border, Side, PatternFill,)
from openpyxl.drawing.image import Image as XLImage

from django.shortcuts import render,redirect,get_object_or_404
from .models import *
from decimal import Decimal
from datetime import date
from timesheet.models import MonthlyTimesheet
from employees.models import Designation
from django.contrib import messages
from .forms import *
from clients.models import Client, ProjectSite
from employees.models import Designation
from num2words import num2words
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Q
from django.core.paginator import Paginator
from django.db.models import Sum

# from django.template.loader import render_to_string
# from weasyprint import HTML
from django.contrib.auth.decorators import login_required
from accounts.decorators import role_required
from django.views.decorators.http import require_POST



# Create your views here.

@login_required
@role_required(["admin", "hr", "accountant"])
def invoice_list(request):

    invoices = Invoice.objects.select_related(
        'company',
        'timesheet',
        'timesheet__project_site',
        'timesheet__project_site__client'
    ).order_by('-invoice_date', '-id')

    search = request.GET.get('search', '')
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')
    status = request.GET.get('status', '')

    if search:
        invoices = invoices.filter(
            Q(invoice_no__icontains=search) |
            Q(timesheet__project_site__name__icontains=search) |
            Q(timesheet__project_site__client__name__icontains=search)
        )

    if month:
        invoices = invoices.filter(
            invoice_date__month=month
        )

    if year:
        invoices = invoices.filter(
            invoice_date__year=year
        )

    if status:
        invoices = invoices.filter(
            status=status
        )

    total_invoices = invoices.count()

    draft_count = invoices.filter(
        status='DRAFT'
    ).count()

    approved_count = invoices.filter(
        status='APPROVED'
    ).count()

    paid_count = invoices.filter(
        status='PAID'
    ).count()  

    paginator = Paginator(invoices, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)  

    return render(request, 'invoice/invoice_list.html', {
        'invoices': page_obj,
        'page_obj': page_obj,
        'search': search,
        'month': month,
        'year': year,
        'status': status,
        'total_invoices':  total_invoices,
        'draft_count': draft_count,
        'approved_count': approved_count,
        'paid_count': paid_count
        
    })



@login_required
@role_required(["admin", "hr", "accountant"])
def billing_rate_list(request):

    client_id = request.GET.get('client')
    project_id = request.GET.get('project')
    designation_id = request.GET.get('designation')
    search = request.GET.get('search', '')

    rates = ProjectBillingRate.objects.select_related(
        'project_site',
        'project_site__client',
        'billing_designation'
    ).order_by(
        'project_site__client__name',
        'project_site__name',
        'billing_designation__name'
    )

    if client_id:
        rates = rates.filter(project_site__client_id=client_id)

    if project_id:
        rates = rates.filter(project_site_id=project_id)

    if designation_id:
        rates = rates.filter(billing_designation_id=designation_id)

    if search:
        rates = rates.filter(
            project_site__name__icontains=search
        )

    clients = Client.objects.all().order_by('name')
    projects = ProjectSite.objects.all().order_by('name')
    designations = Designation.objects.all().order_by('name')

    return render(request, 'invoice/billing_rate_list.html', {
        'rates': rates,
        'clients': clients,
        'projects': projects,
        'designations': designations,
        'client_id': client_id,
        'project_id': project_id,
        'designation_id': designation_id,
        'search': search,
    })


@login_required
@role_required(["admin", "hr", "accountant"])
def billing_rate_create(request):

    if request.method == 'POST':

        form = ProjectBillingRateForm(request.POST)

        if form.is_valid():
            form.save()

            messages.success(
                request,
                'Billing rate added successfully.'
            )

            return redirect('billing_rate_list')

    else:
        form = ProjectBillingRateForm()

    return render(
        request,
        'invoice/billing_rate_form.html',
        {'form': form}
    )

@login_required
@role_required(["admin", "hr", "accountant"])
def billing_rate_detail(request, pk):

    rate = get_object_or_404(ProjectBillingRate, pk=pk)

    return render(request, 'invoice/billing_rate_detail.html', {
        'rate': rate
    })

@login_required
@role_required(["admin", "hr", "accountant"])
def billing_rate_edit(request, pk):

    rate = get_object_or_404(ProjectBillingRate, pk=pk)

    if request.method == 'POST':
        form = ProjectBillingRateForm(request.POST, instance=rate)

        if form.is_valid():
            form.save()
            messages.success(request, 'Billing rate updated successfully.')
            return redirect('billing_rate_detail', pk=rate.id)

    else:
        form = ProjectBillingRateForm(instance=rate)

    return render(request, 'invoice/billing_rate_form.html', {
        'form': form,
        'rate': rate
    })

@login_required
@role_required(["admin", "hr", "accountant"])
def billing_rate_delete(request, pk):

    rate = get_object_or_404(ProjectBillingRate, pk=pk)

    if request.method == 'POST':
        rate.delete()
        messages.success(request, 'Billing rate deleted successfully.')
        return redirect('billing_rate_list')

    return render(request, 'invoice/billing_rate_confirm_delete.html', {
        'rate': rate
    })

@login_required
@role_required(["admin", "hr", "accountant"])
def generate_invoice(request, timesheet_id):

    timesheet = get_object_or_404(
        MonthlyTimesheet,
        id=timesheet_id
    )

    project_site = timesheet.project_site
    project_site.refresh_from_db()

    if timesheet.status != 'APPROVED':
        messages.error(request, 'Only approved timesheets can generate invoices.')
        return redirect('timesheet_detail', pk=timesheet.id)

    if request.method == 'POST':

        form = GenerateInvoiceForm(request.POST)

        if form.is_valid():

            invoice_no = form.cleaned_data['invoice_no']
            invoice_date = form.cleaned_data['invoice_date']
            lpo_no = form.cleaned_data['lpo_no']
            payment_terms = form.cleaned_data['payment_terms']
            description = form.cleaned_data['description']
            terms_conditions = form.cleaned_data['terms_conditions']
            checked_by = form.cleaned_data['checked_by']
            checked_by_designation = form.cleaned_data['checked_by_designation']
            approved_by = form.cleaned_data['approved_by']
            approved_by_designation = form.cleaned_data['approved_by_designation']
            remarks = form.cleaned_data['remarks']

            if Invoice.objects.filter(timesheet=timesheet).exists():
                messages.warning(request, 'Invoice already generated for this timesheet.')
                invoice = Invoice.objects.get(timesheet=timesheet)
                return redirect('invoice_detail', pk=invoice.id)

            if Invoice.objects.filter(invoice_no=invoice_no).exists():
                messages.error(request, 'This invoice number already exists.')
                return redirect('generate_invoice', timesheet_id=timesheet.id)

            grouped_data = {}

            for row in timesheet.rows.all():

                designation_name = row.designation

                if designation_name not in grouped_data:
                    grouped_data[designation_name] = {
                        'qty': 0,
                        'hours': Decimal('0.00')
                    }

                grouped_data[designation_name]['qty'] += 1
                grouped_data[designation_name]['hours'] += Decimal(str(row.total_hours))

            missing_designations = []

            for designation_name in grouped_data.keys():

                designation = Designation.objects.filter(
                    name=designation_name
                ).first()

                if not designation:
                    missing_designations.append(designation_name)
                    continue

                rate_exists = ProjectBillingRate.objects.filter(
                    project_site=project_site,
                    billing_designation=designation
                ).exists()

                if not rate_exists:
                    missing_designations.append(designation_name)

            if missing_designations:
                messages.error(
                    request,
                    'Billing rates missing for: ' + ', '.join(missing_designations)
                )
                return redirect('billing_rate_list')

            invoice = Invoice.objects.create(
                invoice_no=invoice_no,
                company=timesheet.company,
                timesheet=timesheet,
                invoice_date=invoice_date,
                lpo_no=lpo_no,
                payment_terms=payment_terms,
                description=description,
                terms_conditions=terms_conditions,
                checked_by=checked_by,
                checked_by_designation=checked_by_designation,
                approved_by=approved_by,
                approved_by_designation=approved_by_designation,
                remarks=remarks,
                status='DRAFT'
            )

            for designation_name, data in grouped_data.items():

                designation = Designation.objects.get(
                    name=designation_name
                )

                billing_rate = ProjectBillingRate.objects.get(
                    project_site=project_site,
                    billing_designation=designation
                )

                taxable_amount = data['hours'] * billing_rate.hourly_rate
                vat_amount = taxable_amount * Decimal('0.05')
                net_amount = taxable_amount + vat_amount

                InvoiceItem.objects.create(
                    invoice=invoice,
                    designation=designation,
                    qty=data['qty'],
                    total_hours=data['hours'],
                    unit_price=billing_rate.hourly_rate,
                    taxable_amount=taxable_amount,
                    vat_amount=vat_amount,
                    net_amount=net_amount
                )

            messages.success(request, 'Invoice generated successfully.')
            return redirect('invoice_detail', pk=invoice.id)

    else:

        month_name = timesheet.get_month_display()

        default_description = (
            f"Being the charge of contract workers "
            f"for the month of {month_name}-{timesheet.year} "
            f"at {project_site.name}"
        )

        form = GenerateInvoiceForm(initial={
            'lpo_no': project_site.lpo_number or '',
            'payment_terms': '60 Days',
            'description': default_description,
            'terms_conditions':
                "1. Please advise us of any discrepancies within 7 days of the date of invoice.\n"
                "2. Payment should be in the name of HAYAAT MARINE SERVICES.\n"
                "3. Your earliest payment would be highly appreciated.",
            'checked_by': 'Shaji Lawrence',
            'checked_by_designation': 'Accountant',
            'approved_by': 'Lenin Lopez',
            'approved_by_designation': 'Operation Manager',
        })

    return render(request, 'invoice/generate_invoice_form.html', {
        'form': form,
        'timesheet': timesheet
    })

@login_required
@role_required(["admin", "hr", "accountant"])
def invoice_detail(request, pk):

    invoice = get_object_or_404(
        Invoice,
        pk=pk
    )

    items = invoice.items.all()

    taxable_total = sum(
        item.taxable_amount
        for item in items
    )

    vat_total = sum(
        item.vat_amount
        for item in items
    )

    net_total = sum(
        item.net_amount
        for item in items
    )

    return render(
        request,
        'invoice/invoice_detail.html',
        {
            'invoice': invoice,
            'items': items,
            'taxable_total': taxable_total,
            'vat_total': vat_total,
            'net_total': net_total,
        }
    )


@login_required
@role_required(["admin", "hr", "accountant"])
def invoice_print(request, pk):

    invoice = get_object_or_404(
        Invoice.objects.select_related(
            "company",
            "timesheet",
            "timesheet__project_site",
            "timesheet__project_site__client",
        ),
        pk=pk,
    )

    items = invoice.items.select_related("designation").all()

    # Maintain a fixed invoice-table appearance.
    blank_rows = range(max(0, 7 - items.count()))

    total_qty = sum(item.qty for item in items)
    total_hours = sum(item.total_hours for item in items)
    taxable_total = sum(item.taxable_amount for item in items)
    vat_total = sum(item.vat_amount for item in items)
    net_total = sum(item.net_amount for item in items)

    whole = int(net_total)
    fils = round((float(net_total) - whole) * 100)

    amount_words = (
        f"{num2words(whole).title()} Dirhams "
        f"And {num2words(fils).title()} Fils Only"
    )

    return render(
        request,
        "invoice/invoice_print.html",
        {
            "invoice": invoice,
            "items": items,
            "blank_rows": blank_rows,
            "total_qty": total_qty,
            "total_hours": total_hours,
            "taxable_total": taxable_total,
            "vat_total": vat_total,
            "net_total": net_total,
            "amount_words": amount_words,
        },
    )

@login_required
@role_required(["admin", "hr", "accountant"])
def invoice_edit(request, pk):

    invoice = get_object_or_404(Invoice, pk=pk)

    if invoice.status != 'DRAFT':
        messages.error(request, 'Only draft invoices can be edited.')
        return redirect('invoice_detail', pk=invoice.id)

    if request.method == 'POST':
        form = InvoiceEditForm(request.POST, instance=invoice)

        if form.is_valid():
            form.save()
            messages.success(request, 'Invoice updated successfully.')
            return redirect('invoice_detail', pk=invoice.id)

    else:
        form = InvoiceEditForm(instance=invoice)

    return render(request, 'invoice/invoice_form.html', {
        'form': form,
        'invoice': invoice
    })

@login_required
@role_required(["admin", "hr", "accountant"])
def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    user_role = getattr(
        getattr(request.user, "userprofile", None),
        "role",
        None,
    )

    is_admin = (
        request.user.is_superuser
        or user_role == "admin"
    )

    if invoice.status == "PAID":
        messages.error(
            request,
            "A paid invoice cannot be deleted.",
        )
        return redirect(
            "invoice_detail",
            pk=invoice.pk,
        )

    if invoice.status == "APPROVED" and not is_admin:
        messages.error(
            request,
            "Only an administrator can delete an approved invoice.",
        )
        return redirect(
            "invoice_detail",
            pk=invoice.pk,
        )

    if request.method == "POST":
        invoice.delete()

        messages.success(
            request,
            "Invoice deleted successfully.",
        )
        return redirect("invoice_list")

    return render(
        request,
        "invoice/invoice_confirm_delete.html",
        {"invoice": invoice},
    )



@login_required
@role_required(["admin", "hr", "accountant"])
def invoice_export_excel(request, pk):

    invoice = get_object_or_404(
        Invoice.objects.select_related(
            "company",
            "timesheet",
            "timesheet__project_site",
            "timesheet__project_site__client",
        ),
        pk=pk,
    )

    items = list(
        invoice.items.select_related("designation").all()
    )

    company = invoice.company
    project = invoice.timesheet.project_site
    client = project.client

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tax Invoice"

    # Show normal Excel worksheet
    worksheet.sheet_view.showGridLines = True
    worksheet.sheet_view.showRowColHeaders = True
    worksheet.sheet_view.zoomScale = 85

    # Print only invoice borders, not worksheet gridlines
    worksheet.print_options.gridLines = False
    worksheet.print_options.headings = False

    # A4 printing
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1

    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins.left = 0.40
    worksheet.page_margins.right = 0.40
    worksheet.page_margins.top = 0.40
    worksheet.page_margins.bottom = 0.40
    worksheet.page_margins.header = 0
    worksheet.page_margins.footer = 0

    worksheet.print_options.horizontalCentered = True

    # Border and colours
    thin_side = Side(
        style="thin",
        color="000000",
    )

    medium_side = Side(
        style="medium",
        color="000000",
    )

    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    medium_border = Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=medium_side,
    )

    heading_fill = PatternFill(
        fill_type="solid",
        fgColor="D9E2F3",
    )

    # Excel column widths
    column_widths = {
        "A": 3,
        "B": 6,
        "C": 6,
        "D": 17,
        "E": 10,
        "F": 8,
        "G": 11,
        "H": 11,
        "I": 13,
        "J": 10,
        "K": 13,
        "L": 3,
        "M": 10,
        "N": 10,
        "O": 10,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    def style_range(
        cell_range,
        border=None,
        font=None,
        alignment=None,
        fill=None,
    ):
        for row_cells in worksheet[cell_range]:
            for cell in row_cells:

                if border:
                    cell.border = border

                if font:
                    cell.font = font

                if alignment:
                    cell.alignment = alignment

                if fill:
                    cell.fill = fill

    def merge_write(
        cell_range,
        value="",
        border=None,
        font=None,
        alignment=None,
        fill=None,
    ):
        worksheet.merge_cells(cell_range)

        first_cell = worksheet[
            cell_range.split(":")[0]
        ]

        first_cell.value = value

        style_range(
            cell_range,
            border=border,
            font=font,
            alignment=alignment,
            fill=fill,
        )

        return first_cell

    # Leave normal Excel rows above the invoice
    for row_number in range(1, 10):
        worksheet.row_dimensions[row_number].height = 15

    # =========================================================
    # TAX INVOICE title
    # =========================================================

    worksheet.row_dimensions[10].height = 23

    merge_write(
        "B10:K10",
        "TAX INVOICE",
        font=Font(
            name="Times New Roman",
            size=14,
            bold=True,
            underline="single",
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
    )

    # =========================================================
    # Bill-to and invoice information
    # =========================================================

    bill_to_text = (
        "BILL TO:\n"
        f"{client.name.upper()}\n"
        f"{client.address or ''}\n"
        f"Tel: {client.phone or ''}\n"
        f"Email: {client.email or ''}"
    )

    merge_write(
        "B12:F18",
        bill_to_text,
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
        ),
        alignment=Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        ),
    )

    invoice_information = [
        ("INVOICE NO", invoice.invoice_no),
        (
            "DATE",
            invoice.invoice_date.strftime("%d/%m/%Y"),
        ),
        ("LPO NO", invoice.lpo_no or "-"),
        ("SITE", project.name.upper()),
        (
            "PAYMENT TERMS",
            invoice.payment_terms or "-",
        ),
    ]

    information_row = 12

    for label, value in invoice_information:

        merge_write(
            f"G{information_row}:H{information_row}",
            label,
            border=thin_border,
            font=Font(
                name="Times New Roman",
                size=9,
                bold=True,
            ),
            alignment=Alignment(
                horizontal="left",
                vertical="center",
            ),
        )

        merge_write(
            f"I{information_row}:K{information_row}",
            value,
            border=thin_border,
            font=Font(
                name="Times New Roman",
                size=9,
            ),
            alignment=Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            ),
        )

        worksheet.row_dimensions[
            information_row
        ].height = 18

        information_row += 1

    # Blank rows at bottom of information block
    for row_number in range(information_row, 19):

        merge_write(
            f"G{row_number}:H{row_number}",
            "",
            border=thin_border,
        )

        merge_write(
            f"I{row_number}:K{row_number}",
            "",
            border=thin_border,
        )

        worksheet.row_dimensions[
            row_number
        ].height = 18

    # =========================================================
    # TRN information
    # =========================================================

    worksheet.row_dimensions[20].height = 19

    merge_write(
        "B20:F20",
        f"TRN: {client.trn_number or '-'}",
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
    )

    merge_write(
        "G20:K20",
        f"TRN: {company.trn_number or '-'}",
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
    )

    # =========================================================
    # Invoice description
    # =========================================================

    description = invoice.description

    if not description:
        description = (
            "Being the charge of contract workers "
            f"for the month of "
            f"{invoice.timesheet.month}/"
            f"{invoice.timesheet.year}"
        )

    worksheet.row_dimensions[22].height = 16
    worksheet.row_dimensions[23].height = 16

    merge_write(
        "B22:K23",
        description,
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        ),
    )

    # =========================================================
    # Invoice item headings
    # =========================================================

    header_row = 25

    worksheet.row_dimensions[header_row].height = 35

    merge_write(
        "B25:C25",
        "SL NO.",
        border=medium_border,
        fill=heading_fill,
    )

    merge_write(
        "D25:E25",
        "PARTICULAR",
        border=medium_border,
        fill=heading_fill,
    )

    worksheet["F25"] = "QTY"
    worksheet["G25"] = "TOTAL\nHOURS"
    worksheet["H25"] = "UNIT\nPRICE"
    worksheet["I25"] = "TAXABLE\nAMOUNT"
    worksheet["J25"] = "5% VAT"
    worksheet["K25"] = "NET\nAMOUNT"

    style_range(
        "B25:K25",
        border=medium_border,
        fill=heading_fill,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        ),
    )

    # =========================================================
    # Invoice items
    # =========================================================

    first_item_row = 26
    item_row = first_item_row

    total_qty = 0
    total_hours = Decimal("0.00")
    taxable_total = Decimal("0.00")
    vat_total = Decimal("0.00")
    net_total = Decimal("0.00")

    for index, item in enumerate(items, start=1):

        merge_write(
            f"B{item_row}:C{item_row}",
            index,
            border=thin_border,
        )

        merge_write(
            f"D{item_row}:E{item_row}",
            item.designation.name.upper(),
            border=thin_border,
        )

        worksheet.cell(
            row=item_row,
            column=6,
            value=item.qty,
        )

        worksheet.cell(
            row=item_row,
            column=7,
            value=float(item.total_hours),
        )

        worksheet.cell(
            row=item_row,
            column=8,
            value=float(item.unit_price),
        )

        worksheet.cell(
            row=item_row,
            column=9,
            value=float(item.taxable_amount),
        )

        worksheet.cell(
            row=item_row,
            column=10,
            value=float(item.vat_amount),
        )

        worksheet.cell(
            row=item_row,
            column=11,
            value=float(item.net_amount),
        )

        style_range(
            f"B{item_row}:K{item_row}",
            border=thin_border,
            font=Font(
                name="Times New Roman",
                size=9,
            ),
            alignment=Alignment(
                horizontal="center",
                vertical="center",
            ),
        )

        worksheet.cell(
            row=item_row,
            column=4,
        ).alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        for column_number in range(7, 12):
            worksheet.cell(
                row=item_row,
                column=column_number,
            ).number_format = "0.00"

        worksheet.row_dimensions[item_row].height = 19

        total_qty += item.qty
        total_hours += item.total_hours
        taxable_total += item.taxable_amount
        vat_total += item.vat_amount
        net_total += item.net_amount

        item_row += 1

    # Minimum four visible item rows
    item_slots = max(4, len(items))

    while item_row < first_item_row + item_slots:

        merge_write(
            f"B{item_row}:C{item_row}",
            "",
            border=thin_border,
        )

        merge_write(
            f"D{item_row}:E{item_row}",
            "",
            border=thin_border,
        )

        style_range(
            f"B{item_row}:K{item_row}",
            border=thin_border,
        )

        worksheet.row_dimensions[item_row].height = 19

        item_row += 1

    # =========================================================
    # Total row
    # =========================================================

    total_row = item_row

    merge_write(
        f"B{total_row}:C{total_row}",
        "",
        border=medium_border,
    )

    merge_write(
        f"D{total_row}:E{total_row}",
        "TOTAL",
        border=medium_border,
    )

    worksheet.cell(
        row=total_row,
        column=6,
        value=total_qty,
    )

    worksheet.cell(
        row=total_row,
        column=7,
        value=float(total_hours),
    )

    worksheet.cell(
        row=total_row,
        column=9,
        value=float(taxable_total),
    )

    worksheet.cell(
        row=total_row,
        column=10,
        value=float(vat_total),
    )

    worksheet.cell(
        row=total_row,
        column=11,
        value=float(net_total),
    )

    style_range(
        f"B{total_row}:K{total_row}",
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
    )

    for column_number in range(7, 12):
        worksheet.cell(
            row=total_row,
            column=column_number,
        ).number_format = "0.00"

    worksheet.row_dimensions[total_row].height = 21

    # =========================================================
    # Amount in words
    # =========================================================

    amount_row = total_row + 2

    whole = int(net_total)
    fils = round(
        (float(net_total) - whole) * 100
    )

    amount_words = (
        f"{num2words(whole).title()} Dirhams "
        f"And {num2words(fils).title()} Fils Only"
    )

    merge_write(
        f"B{amount_row}:K{amount_row + 1}",
        f"TOTAL (AED) {amount_words}",
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        ),
    )

    worksheet.row_dimensions[amount_row].height = 16
    worksheet.row_dimensions[amount_row + 1].height = 16

    # =========================================================
    # Terms
    # =========================================================

    terms_start = amount_row + 3

    default_terms = [
        (
            "1. Please advise us of any discrepancies "
            "within 7 days of the date of invoice."
        ),
        (
            f"2. Payment should be in the name of "
            f"{company.name.upper()}."
        ),
        (
            "3. Your earliest payment would be "
            "highly appreciated."
        ),
        "",
    ]

    if invoice.terms_conditions:
        supplied_terms = (
            invoice.terms_conditions.splitlines()
        )

        terms = supplied_terms[:4]

        while len(terms) < 4:
            terms.append("")
    else:
        terms = default_terms

    for position, term in enumerate(terms):

        current_row = terms_start + position

        merge_write(
            f"B{current_row}:K{current_row}",
            term,
            border=thin_border,
            font=Font(
                name="Times New Roman",
                size=9,
            ),
            alignment=Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            ),
        )

        worksheet.row_dimensions[current_row].height = 16

    # =========================================================
    # Signatures
    # =========================================================

    signature_title_row = terms_start + 5
    signature_start_row = signature_title_row + 1
    signature_end_row = signature_start_row + 4

    merge_write(
        f"B{signature_title_row}:D{signature_title_row}",
        "Checked by",
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
    )

    merge_write(
        f"E{signature_title_row}:G{signature_title_row}",
        "Approved by",
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
    )

    merge_write(
        f"H{signature_title_row}:K{signature_title_row}",
        "Received by",
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
    )

    merge_write(
        f"B{signature_start_row}:D{signature_end_row}",
        (
            "\n\n"
            f"{invoice.checked_by or ''}\n"
            f"{invoice.checked_by_designation or ''}"
        ),
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="bottom",
            wrap_text=True,
        ),
    )

    merge_write(
        f"E{signature_start_row}:G{signature_end_row}",
        (
            "\n\n"
            f"{invoice.approved_by or ''}\n"
            f"{invoice.approved_by_designation or ''}"
        ),
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
        ),
        alignment=Alignment(
            horizontal="center",
            vertical="bottom",
            wrap_text=True,
        ),
    )

    merge_write(
        f"H{signature_start_row}:K{signature_end_row}",
        (
            "\n\n"
            "Name & Signature: ____________________\n"
            "Date: ______________________________"
        ),
        border=medium_border,
        font=Font(
            name="Times New Roman",
            size=9,
        ),
        alignment=Alignment(
            horizontal="left",
            vertical="bottom",
            wrap_text=True,
        ),
    )

    for row_number in range(
        signature_start_row,
        signature_end_row + 1,
    ):
        worksheet.row_dimensions[row_number].height = 16

    # Print area: no logo and no company footer
    worksheet.print_area = (
        f"B10:K{signature_end_row}"
    )

    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1

    # Select Tax Invoice when opened
    worksheet.sheet_view.selection[0].activeCell = "B10"
    worksheet.sheet_view.selection[0].sqref = "B10"

    safe_invoice_number = (
        invoice.invoice_no
        .replace("/", "-")
        .replace("\\", "-")
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="'
        f'{safe_invoice_number}.xlsx"'
    )

    workbook.save(response)

    return response

@login_required
@role_required(["admin", "hr", "accountant"])
@require_POST
def invoice_approve(request, pk):

    invoice = get_object_or_404(
        Invoice,
        pk=pk
    )

    if invoice.status != 'DRAFT':
        messages.error(
            request,
            'Only draft invoices can be approved.'
        )
        return redirect(
            'invoice_detail',
            pk=invoice.id
        )

    invoice.status = 'APPROVED'
    invoice.save()

    messages.success(
        request,
        'Invoice approved successfully.'
    )

    return redirect(
        'invoice_detail',
        pk=invoice.id
    )



@login_required
@role_required(["admin", "hr", "accountant"])
@require_POST
def invoice_mark_paid(request, pk):

    invoice = get_object_or_404(Invoice, pk=pk)

    if invoice.status != 'APPROVED':
        messages.error(
            request,
            'Only approved invoices can be marked as paid.'
        )
        return redirect('invoice_detail', pk=invoice.id)

    invoice.status = 'PAID'
    invoice.paid_date = date.today()
    invoice.save()

    messages.success(request, 'Invoice marked as paid.')

    return redirect('invoice_detail', pk=invoice.id)



# def invoice_pdf(request, pk):

#     invoice = get_object_or_404(
#         Invoice,
#         pk=pk
#     )

#     items = invoice.items.all()

#     taxable_total = sum(
#         item.taxable_amount
#         for item in items
#     )

#     vat_total = sum(
#         item.vat_amount
#         for item in items
#     )

#     net_total = sum(
#         item.net_amount
#         for item in items
#     )

#     html_string = render_to_string(
#         'invoice/invoice_print.html',
#         {
#             'invoice': invoice,
#             'items': items,
#             'taxable_total': taxable_total,
#             'vat_total': vat_total,
#             'net_total': net_total,
#         }
#     )

#     pdf = HTML(
#         string=html_string,
#         base_url=request.build_absolute_uri('/')
#     ).write_pdf()

#     response = HttpResponse(
#         pdf,
#         content_type='application/pdf'
#     )

#     response['Content-Disposition'] = (
#         f'attachment; filename="{invoice.invoice_no}.pdf"'
#     )

#     return response


@login_required
@role_required(["admin", "hr", "accountant"])
def outstanding_invoice_report(request):

    search = request.GET.get('search', '')
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')

    invoices = Invoice.objects.select_related(
        'company',
        'timesheet',
        'timesheet__project_site',
        'timesheet__project_site__client'
    ).filter(
        status='APPROVED'
    ).order_by('invoice_date')

    if search:
        invoices = invoices.filter(
            Q(invoice_no__icontains=search) |
            Q(timesheet__project_site__name__icontains=search) |
            Q(timesheet__project_site__client__name__icontains=search)
        )

    if month:
        invoices = invoices.filter(invoice_date__month=month)

    if year:
        invoices = invoices.filter(invoice_date__year=year)

    today = date.today()

    report_data = []
    total_outstanding = 0
    total_days = 0

    for invoice in invoices:
        net_total = sum(item.net_amount for item in invoice.items.all())
        days_outstanding = (today - invoice.invoice_date).days

        total_outstanding += net_total
        total_days += days_outstanding

        report_data.append({
            'invoice': invoice,
            'net_total': net_total,
            'days_outstanding': days_outstanding,
        })

    outstanding_count = len(report_data)

    average_days = 0
    if outstanding_count > 0:
        average_days = total_days // outstanding_count

    return render(request, 'invoice/outstanding_invoice_report.html', {
        'report_data': report_data,
        'total_outstanding': total_outstanding,
        'outstanding_count': outstanding_count,
        'average_days': average_days,
        'search': search,
        'month': month,
        'year': year,
    })



@login_required
@role_required(["admin", "hr", "accountant"])
def payment_collection_report(request):

    search = request.GET.get('search', '')
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')

    show_cards = False

    if search or month or year:
        show_cards = True

    invoices = Invoice.objects.select_related(
        'company',
        'timesheet',
        'timesheet__project_site',
        'timesheet__project_site__client'
    ).filter(
        status='PAID'
    ).order_by('-paid_date', '-invoice_date')

    if search:
        invoices = invoices.filter(
            Q(invoice_no__icontains=search) |
            Q(timesheet__project_site__name__icontains=search) |
            Q(timesheet__project_site__client__name__icontains=search)
        )

    if month:
        invoices = invoices.filter(paid_date__month=month)

    if year:
        invoices = invoices.filter(paid_date__year=year)

    report_data = []
    total_collected = 0

    for invoice in invoices:

        net_total = sum(
            item.net_amount
            for item in invoice.items.all()
        )

        total_collected += net_total

        report_data.append({
            'invoice': invoice,
            'net_total': net_total,
        })

    collected_count = len(report_data)

    average_collection = 0

    if collected_count > 0:
        average_collection = total_collected / collected_count

    return render(request, 'invoice/payment_collection_report.html', {
        'report_data': report_data,
        'total_collected': total_collected,
        'collected_count': collected_count,
        'average_collection': average_collection,
        'show_cards': show_cards,
        'search': search,
        'month': month,
        'year': year,
    })




@login_required
@role_required(["admin"])
@require_POST
def invoice_reopen(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    if invoice.status != "APPROVED":
        messages.error(
            request,
            "Only approved, unpaid invoices can be reopened.",
        )
        return redirect("invoice_detail", pk=invoice.pk)

    invoice.status = "DRAFT"
    invoice.save(update_fields=["status"])

    messages.success(
        request,
        "Invoice reopened successfully.",
    )

    return redirect("invoice_detail", pk=invoice.pk)