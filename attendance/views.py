from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from decimal import Decimal
from .models import *
from clients.models import ProjectSite, EmployeeAssignment
from django.db.models import Q
from django.db.models import Sum, Count
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.decorators import login_required
from accounts.decorators import role_required


# Create your views here.
@login_required
@role_required(["admin", "hr", "accountant"])
def attendance_entry(request):

    projects = ProjectSite.objects.filter(is_active=True).order_by('name')
    
    selected_project = None
    selected_date = None
    working_hours = None
    assignments = []
    show_no_employee_message = False

    if request.method == 'POST':

        project_id = request.POST.get('project_site')
        selected_date = request.POST.get('attendance_date')
        working_hours = Decimal(request.POST.get('working_hours') or 0)

        selected_project = ProjectSite.objects.get(id=project_id)

        attendance_exists = DailyAttendance.objects.filter(
        project_site=selected_project,
        attendance_date=selected_date
        ).exists()

        if attendance_exists and 'save_attendance' not in request.POST:
            messages.warning(
                request,
                'Attendance already exists for this project/site and date. Please edit from Attendance List.'
            )
            return redirect(
                f'/attendance/list/?project={selected_project.id}&from_date={selected_date}&to_date={selected_date}'
            )
        assignments = EmployeeAssignment.objects.select_related(
            'employee',
            'designation',
            'project_site'
        ).filter(
            project_site=selected_project,
            status='ACTIVE'
        )
        if not assignments.exists() and 'save_attendance' not in request.POST:messages.warning(
          request,'No active employees are assigned to this project/site. Please assign employees first.'
        )


        if 'save_attendance' in request.POST:

            for assignment in assignments:

                status = request.POST.get(f'status_{assignment.id}')
                hours = request.POST.get(f'hours_{assignment.id}')
                remarks = request.POST.get(f'remarks_{assignment.id}', '')

                if status in ['ABSENT', 'LEAVE', 'HOLIDAY']:
                    hours = 0

                DailyAttendance.objects.update_or_create(
                    attendance_date=selected_date,
                    project_site=selected_project,
                    employee=assignment.employee,
                    defaults={
                        'assignment': assignment,
                        'status': status,
                        'total_hours': hours,
                        'remarks': remarks,
                    }
                )

            messages.success(request, 'Attendance saved successfully.')
            return redirect('attendance_entry')

    context = {
        'projects': projects,
        'selected_project': selected_project,
        'selected_date': selected_date,
        'working_hours': working_hours,
        'assignments': assignments,
        'show_no_employee_message': show_no_employee_message
    }

    return render(request, 'attendance/attendance_entry.html', context)

@login_required
@role_required(["admin", "hr", "accountant"])
def attendance_list(request):

    projects = ProjectSite.objects.filter(is_active=True)

    filters_applied = False

    project_id = request.GET.get('project')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    status = request.GET.get('status')
    employee = request.GET.get('employee')

    attendances = DailyAttendance.objects.select_related(
        'employee',
        'project_site',
        'project_site__client',
        'assignment',
        'assignment__designation'
    ).order_by('attendance_date', 'employee__first_name')

    if project_id:
        filters_applied = True
        attendances = attendances.filter(project_site_id=project_id)

    if from_date:
        filters_applied = True
        attendances = attendances.filter(attendance_date__gte=from_date)

    if to_date:
        filters_applied = True
        attendances = attendances.filter(attendance_date__lte=to_date)

    if status:
        filters_applied = True
        attendances = attendances.filter(status=status)

    if employee:
        filters_applied = True
        attendances = attendances.filter(
            Q(employee__employee_id__icontains=employee) |
            Q(employee__first_name__icontains=employee) |
            Q(employee__last_name__icontains=employee)
        )


    if from_date or to_date:
      attendances = attendances.order_by(
        'attendance_date',
        'employee__first_name',
        'employee__last_name'
    )
    else:
        attendances = attendances.order_by(
            '-attendance_date',
            '-id'
    )

    if filters_applied:
        present_count = attendances.filter(status='PRESENT').count()
        absent_count = attendances.filter(status='ABSENT').count()
        leave_count = attendances.filter(status='LEAVE').count()
        holiday_count = attendances.filter(status='HOLIDAY').count()

        total_hours = attendances.aggregate(
            total=Sum('total_hours')
        )['total'] or 0
    else:
        present_count = 0
        absent_count = 0
        leave_count = 0
        holiday_count = 0
        total_hours = 0

    context = {
        'attendances': attendances,
        'projects': projects,
        'present_count': present_count,
        'absent_count': absent_count,
        'leave_count': leave_count,
        'holiday_count': holiday_count,
        'total_hours': total_hours,
    }

    return render(
        request,
        'attendance/attendance_list.html',
        context
    )


@login_required
@role_required(["admin", "hr", "accountant"])
def attendance_update(request, pk):

    attendance = get_object_or_404(
        DailyAttendance,
        pk=pk
    )

    if request.method == 'POST':

        attendance.status = request.POST.get('status')
        attendance.total_hours = request.POST.get('total_hours')
        attendance.remarks = request.POST.get('remarks')

        attendance.save()

        messages.success(
            request,
            'Attendance updated successfully.'
        )

        return redirect('attendance_list')

    return render(
        request,
        'attendance/attendance_form.html',
        {
            'attendance': attendance
        }
    )

@login_required
@role_required(["admin", "hr", "accountant"])
def attendance_delete(request, pk):

    attendance = get_object_or_404(
        DailyAttendance,
        pk=pk
    )

    if request.method == 'POST':
        attendance.delete()

        messages.success(
            request,
            'Attendance deleted successfully.'
        )

        return redirect('attendance_list')

    return render(
        request,
        'attendance/attendance_confirm_delete.html',
        {
            'attendance': attendance
        }
    )


@login_required
@role_required(["admin", "hr", "accountant"])
def export_attendance_excel(request):

    attendances = DailyAttendance.objects.select_related(
        'employee',
        'assignment',
        'assignment__designation',
        'project_site',
        'project_site__client'
    ).order_by('attendance_date','employee__first_name')

    project_id = request.GET.get('project')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    status = request.GET.get('status')
    employee = request.GET.get('employee')

    if project_id:
        attendances = attendances.filter(project_site_id=project_id)

    if from_date:
        attendances = attendances.filter(attendance_date__lte=to_date)

    if to_date:
        attendances = attendances.filter(attendance_date__gte=from_date)

    if status:
        attendances = attendances.filter(status=status)

    if employee:
        attendances = attendances.filter(
            Q(employee__employee_id__icontains=employee) |
            Q(employee__first_name__icontains=employee) |
            Q(employee__last_name__icontains=employee)
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Attendance Report"

    headers = [
        'Date',
        'Employee ID',
        'Employee Name',
        'Designation',
        'Client',
        'Project Site',
        'Status',
        'Hours',
        'Remarks'
    ]

    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    row_num = 2

    for attendance in attendances:
        worksheet.cell(row=row_num, column=1, value=str(attendance.attendance_date))
        worksheet.cell(row=row_num, column=2, value=attendance.employee.employee_id)
        worksheet.cell(
            row=row_num,
            column=3,
            value=f"{attendance.employee.first_name} {attendance.employee.last_name}"
        )
        worksheet.cell(
            row=row_num,
            column=4,
            value=attendance.assignment.designation.name if attendance.assignment.designation else ''
        )
        worksheet.cell(row=row_num, column=5, value=attendance.project_site.client.name)
        worksheet.cell(row=row_num, column=6, value=attendance.project_site.name)
        worksheet.cell(row=row_num, column=7, value=attendance.status)
        worksheet.cell(row=row_num, column=8, value=float(attendance.total_hours))
        worksheet.cell(row=row_num, column=9, value=attendance.remarks)

        row_num += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'

    workbook.save(response)

    return response


@login_required
@role_required(["admin", "hr", "accountant"])
def attendance_bulk_edit(request):

    project_id = request.GET.get('project')
    attendance_date = request.GET.get('date')

    attendances = DailyAttendance.objects.select_related(
        'employee',
        'assignment',
        'assignment__designation',
        'project_site'
    ).filter(
        project_site_id=project_id,
        attendance_date=attendance_date
    ).order_by(
        'employee__first_name',
        'employee__last_name'
    )

    if request.method == 'POST':

        for attendance in attendances:

            delete_checked = request.POST.get(f'delete_{attendance.id}')

            if delete_checked:
                attendance.delete()
                continue

            attendance.status = request.POST.get(f'status_{attendance.id}')
            attendance.total_hours = request.POST.get(f'hours_{attendance.id}')
            attendance.remarks = request.POST.get(f'remarks_{attendance.id}', '')

            if attendance.status in ['ABSENT', 'LEAVE', 'HOLIDAY']:
                attendance.total_hours = 0

            attendance.save()

        messages.success(request, 'Attendance updated successfully.')

        return redirect(
            f'/attendance/list/?project={project_id}&from_date={attendance_date}&to_date={attendance_date}'
        )

    return render(request, 'attendance/attendance_bulk_edit.html', {
        'attendances': attendances,
        'project_id': project_id,
        'attendance_date': attendance_date,
    })

